from openpyxl import load_workbook
import os


def get_login_data():

    current = os.path.dirname(__file__)

    file = os.path.join(
        current,
        "..",
        "testdata",
        "login_data.xlsx"
    )

    wb = load_workbook(file)

    sheet = wb.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data